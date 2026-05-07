"""
스크리닝 데이터 API

daily screening xlsx 파일 조회 → 추후 DB 전환 염두한 구조

파일 기반 소스와 DB 기반 소스를 추상화하여,
데이터 소스 교체가 용이하도록 설계.
"""

import json
import logging
import os
import re
from datetime import datetime
from pathlib import Path
from typing import Optional

import openpyxl
from fastapi import APIRouter, Depends, HTTPException, Query
from sqlalchemy.orm import Session

from ..database import get_reports_db
from ..dependencies import get_user_from_token
from ..models import User
from ..settings import get_settings

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/screening", tags=["Screening"])

# ── 파일 패턴 ─────────────────────────────────────────────
# KR_stock_screening_YYMMDD.xlsx  (ex: KR_stock_screening_260507.xlsx)
FILENAME_PATTERN = re.compile(r"^KR_stock_screening_(\d{6})\.xlsx$")


# ── 설정 기반 파일 경로 ────────────────────────────────────
def get_screening_dir() -> str:
    """
    settings.SCREENING_FILES_PATH 를 반환.
    기본값: /home/ubuntu/prod/telegram-stock-info-bot/send
    """
    return get_settings().screening_files_path


# ── 데이터 소스 추상화 (향후 DB 전환 대비) ─────────────────
class ScreeningFileSource:
    """
    일별 xlsx 파일 기반 데이터 소스.
    추후 DBSource로 교체 가능.
    """

    @staticmethod
    def list_files(base_dir: str) -> list[dict]:
        """send 디렉토리에서 일자별 xlsx 파일 목록 반환"""
        if not os.path.isdir(base_dir):
            logger.warning("Screening directory not found: %s", base_dir)
            return []

        files = []
        for fname in sorted(os.listdir(base_dir), reverse=True):
            match = FILENAME_PATTERN.match(fname)
            if not match:
                continue

            yymmdd = match.group(1)
            fpath = os.path.join(base_dir, fname)
            mtime = os.path.getmtime(fpath)
            size = os.path.getsize(fpath)
            # YYMMDD -> YYYY-MM-DD
            try:
                date_obj = datetime.strptime(yymmdd, "%y%m%d")
                date_str = date_obj.strftime("%Y-%m-%d")
            except ValueError:
                date_str = yymmdd

            files.append({
                "filename": fname,
                "date": date_str,
                "dateCode": yymmdd,
                "size": size,
                "sizeLabel": _format_size(size),
                "mtime": datetime.fromtimestamp(mtime).isoformat(),
            })

        return files

    @staticmethod
    def read_file(base_dir: str, filename: str) -> dict:
        """xlsx 파일을 읽어 시트별 JSON 데이터 반환"""
        fpath = os.path.join(base_dir, filename)
        if not os.path.isfile(fpath):
            raise HTTPException(status_code=404, detail=f"File not found: {filename}")

        try:
            wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
        except Exception as e:
            logger.error("Failed to load xlsx %s: %s", filename, e)
            raise HTTPException(status_code=500, detail="Failed to read file")

        result = {
            "filename": filename,
            "sheets": [],
        }

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            header = [str(h) if h is not None else "" for h in rows[0]]
            data = []
            for row in rows[1:]:
                item = {}
                for idx, val in enumerate(row):
                    key = header[idx] if idx < len(header) else f"col_{idx}"
                    # NaN, None 처리
                    if val is None or (isinstance(val, float) and val != val):
                        item[key] = None
                    elif isinstance(val, float):
                        item[key] = val
                    else:
                        item[key] = str(val).strip() if val else None
                data.append(item)

            result["sheets"].append({
                "name": sheet_name,
                "columns": header,
                "rows": len(data),
                "data": data,
            })

        wb.close()
        return result


# ── 데이터 소스 레지스트리 (향후 DB 소스 추가) ────────────
_DATA_SOURCES = {
    "file": ScreeningFileSource(),
    # "db": DBScreeningSource(),  # TODO: DB 마이그레이션 시 추가
}


# ── API Endpoints ──────────────────────────────────────────

@router.get("/files")
@router.get("/files/")
async def list_screening_files(
    current_user: User = Depends(get_user_from_token),
):
    """
    일별 스크리닝 xlsx 파일 목록 조회 (최신순)
    
    Returns:
        [
            {
                "filename": "KR_stock_screening_260507.xlsx",
                "date": "2026-05-07",
                "dateCode": "260507",
                "size": 1234567,
                "sizeLabel": "1.2 MB",
                "mtime": "2026-05-07T10:30:00"
            },
            ...
        ]
    """
    base_dir = get_screening_dir()
    files = ScreeningFileSource.list_files(base_dir)
    return files


@router.get("/files/{filename}")
@router.get("/files/{filename}/")
async def get_screening_file(
    filename: str,
    sheet: Optional[str] = Query(None, description="특정 시트만 조회"),
    limit: Optional[int] = Query(None, description="행 제한 (전체 조회 시 생략)"),
    current_user: User = Depends(get_user_from_token),
):
    """
    특정 일자 스크리닝 파일의 데이터를 JSON으로 조회
    
    Returns:
        {
            "filename": "KR_stock_screening_260507.xlsx",
            "sheets": [
                {
                    "name": "KOSPI",
                    "columns": ["종목코드", "종목명", ...],
                    "rows": 950,
                    "data": [ ... ]
                },
                ...
            ]
        }
    """
    # 파일명 검증
    if not FILENAME_PATTERN.match(filename):
        raise HTTPException(status_code=400, detail="Invalid filename format")

    base_dir = get_screening_dir()
    result = ScreeningFileSource.read_file(base_dir, filename)

    # 시트 필터
    if sheet:
        filtered = [s for s in result["sheets"] if s["name"] == sheet]
        if not filtered:
            raise HTTPException(status_code=404, detail=f"Sheet not found: {sheet}")
        result["sheets"] = filtered

    # 행 제한
    if limit and limit > 0:
        for s in result["sheets"]:
            s["data"] = s["data"][:limit]
            s["rows"] = len(s["data"])

    return result


@router.get("/latest")
@router.get("/latest/")
async def get_latest_screening(
    sheet: Optional[str] = Query(None, description="특정 시트만 조회"),
    limit: Optional[int] = Query(100, description="조회할 행 수 (기본 100)"),
    current_user: User = Depends(get_user_from_token),
):
    """
    가장 최신 스크리닝 파일 데이터 조회 (기본 100행)
    """
    base_dir = get_screening_dir()
    files = ScreeningFileSource.list_files(base_dir)

    if not files:
        raise HTTPException(status_code=404, detail="No screening files found")

    latest = files[0]["filename"]
    result = ScreeningFileSource.read_file(base_dir, latest)

    if sheet:
        filtered = [s for s in result["sheets"] if s["name"] == sheet]
        if not filtered:
            raise HTTPException(status_code=404, detail=f"Sheet not found: {sheet}")
        result["sheets"] = filtered

    if limit and limit > 0:
        for s in result["sheets"]:
            s["data"] = s["data"][:limit]
            s["rows"] = len(s["data"])

    return result


def _format_size(size_bytes: int) -> str:
    if size_bytes >= 1024 * 1024:
        return f"{size_bytes / (1024 * 1024):.1f} MB"
    elif size_bytes >= 1024:
        return f"{size_bytes / 1024:.1f} KB"
    return f"{size_bytes} B"
